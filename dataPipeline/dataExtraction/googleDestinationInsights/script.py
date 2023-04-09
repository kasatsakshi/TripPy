import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl


options = webdriver.ChromeOptions()
options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
chrome_driver_binary = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver_binary, chrome_options=options)
driver.get('https://destinationinsights.withgoogle.com/') 

counter = 1 
listtotal = [] 
time.sleep(50) 
table2 = driver.find_element('xpath',"/html/body/div[1]/div[13]/div[2]/div[4]") 
table2clean = table2.get_attribute("outerHTML") 
table2soup = BeautifulSoup(table2clean, "lxml") 
growthregions = table2soup.find_all("span", {"class": "geographic-demand__name ng-binding"}) 
growthregionsnumber = table2soup.find_all("span", {"class": "geographic-demand__total glue-small-text ng-binding"}) 

growth = [] 
for x in range(len(growthregions)): 
    growth.append([growthregions[x].text,growthregionsnumber[x].text]) 

dictionarytotal = {'USA' : [growth]} 
listtotal.append(dictionarytotal) 

filepath = "Test.xlsx" 
wb = openpyxl.Workbook() 
wb.save(filepath) 
for x in listtotal: 
    for k,v in x.items(): 
        key_dict = k 
        value_dict = v 
        ws1 = wb.create_sheet("Sheet_A") 
        if len(str(k)) < 31: 
            ws1.title = str(k) 
            ws = wb[str(k)] 
        else: 
            ws1.title = str(k)[0:30] 
            ws = wb[str(k)[0:30]] 
        counter = 1 
    print(value_dict)
    for y in value_dict: 
        counter2 = 2 
        ws.cell(row=1, column=1).value = "TOP DEMAND BY DESTINATION CITY" 
        ws.cell(row=1, column=2).value = "Interest" 
        for z in y: 
            ws.cell(row=counter2, column=counter).value = z[0] 
            ws.cell(row=counter2, column=counter+1).value = z[1] 
            counter2 = counter2 + 1 
del wb["Sheet"] 
wb.save(filepath) 
wb.close()